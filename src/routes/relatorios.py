# Imports principais
import io
from collections import defaultdict
from flask import Blueprint, render_template, request, jsonify, send_file, flash, redirect, url_for
from datetime import datetime
from flask_login import login_required, current_user
from src.models import db, Lancamento, Entidade, ContaBanco, FluxoContaModel
from sqlalchemy import func, or_
from datetime import datetime, date
from decimal import Decimal
from types import SimpleNamespace
import logging
try:
	from openpyxl import Workbook
except Exception:
	Workbook = None
	logging.getLogger(__name__).warning("openpyxl not available; Excel exports disabled", exc_info=True)


# Definição do blueprint deve vir logo após os imports principais
relatorios_bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')
def _parse_date_filter(value, field_label):
	if not value:
		return None
	try:
		return datetime.strptime(value, '%Y-%m-%d').date()
	except ValueError:
		flash(f'Data inválida para "{field_label}". Use o formato YYYY-MM-DD.', 'warning')
		return None


def _build_listagem_lancamentos_context(args):
	# Calcular saldo inicial total das contas filtradas
	def get_saldo_inicial_por_conta():
		if conta_fluxo_id:
			contas = ContaBanco.query.filter(
				ContaBanco.empresa_id == current_user.empresa_id,
				ContaBanco.id == conta_fluxo_id
			).all()
		else:
			contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
		return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

	def get_saldo_inicial_total():
		return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

	data_lanc_de = args.get('data_lanc_de', '')
	data_lanc_ate = args.get('data_lanc_ate', '')
	data_venc_de = args.get('data_venc_de', '')
	data_venc_ate = args.get('data_venc_ate', '')
	tipo_movimento = args.get('tipo_movimento', '')
	entidade_tipo = args.get('entidade_tipo', '')
	entidade_id = args.get('entidade_id', type=int)
	conta_fluxo_id = args.get('conta_fluxo_id', type=int)
	status = args.get('status', '')

	data_lanc_de_dt = _parse_date_filter(data_lanc_de, 'Data de lançamento (de)')
	data_lanc_ate_dt = _parse_date_filter(data_lanc_ate, 'Data de lançamento (até)')
	data_venc_de_dt = _parse_date_filter(data_venc_de, 'Data de vencimento (de)')
	data_venc_ate_dt = _parse_date_filter(data_venc_ate, 'Data de vencimento (até)')

	from sqlalchemy.orm import joinedload
	query = Lancamento.query.filter(Lancamento.empresa_id == current_user.empresa_id).options(joinedload(Lancamento.entidade))

	if data_lanc_de_dt:
		query = query.filter(Lancamento.data_evento >= data_lanc_de_dt)
	if data_lanc_ate_dt:
		query = query.filter(Lancamento.data_evento <= data_lanc_ate_dt)
	if data_venc_de_dt:
		query = query.filter(Lancamento.data_vencimento >= data_venc_de_dt)
	if data_venc_ate_dt:
		query = query.filter(Lancamento.data_vencimento <= data_venc_ate_dt)
	if tipo_movimento in ('P', 'R'):
		query = query.filter(Lancamento.fluxo_conta.has(FluxoContaModel.tipo == tipo_movimento))
	if entidade_tipo in ('C', 'F'):
		query = query.filter(Lancamento.entidade.has(Entidade.tipo == entidade_tipo))
	if entidade_id:
		query = query.filter(Lancamento.entidade_id == entidade_id)
	if conta_fluxo_id:
		query = query.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
	if status:
		query = query.filter(Lancamento.status == status)

	lancamentos = query.order_by(
		Lancamento.data_evento.desc(),
		Lancamento.data_vencimento.desc(),
		Lancamento.id.desc()
	).all()

	total_pago = Decimal('0.00')
	total_recebido = Decimal('0.00')
	total_valor_imposto = Decimal('0.00')
	total_valor_outros_custos = Decimal('0.00')
	for lancamento in lancamentos:
		tipo = lancamento.fluxo_conta.tipo if lancamento.fluxo_conta else None
		valor_pago = Decimal(str(lancamento.valor_pago or 0))
		valor_recebido = Decimal(str(lancamento.valor_pago or 0))
		if tipo == 'P':
			total_pago += valor_pago
		elif tipo == 'R':
			total_recebido += valor_recebido
		total_valor_imposto += Decimal(str(lancamento.valor_imposto or 0))
		total_valor_outros_custos += Decimal(str(lancamento.valor_outros_custos or 0))

	# Filtros: se entidade_tipo == 'C' (cliente), só mostra recebidos; se 'F' (fornecedor), só pagos
	mostrar_pago = entidade_tipo != 'C'
	mostrar_recebido = entidade_tipo != 'F'

	entidades = Entidade.query.filter_by(
		empresa_id=current_user.empresa_id,
		ativo=True
	).order_by(Entidade.nome.asc()).all()
	contas_fluxo = FluxoContaModel.query.filter_by(
		empresa_id=current_user.empresa_id,
		ativo=True
	).order_by(FluxoContaModel.codigo.asc()).all()

	return {
		'lancamentos': lancamentos,
		'entidades': entidades,
		'contas_fluxo': contas_fluxo,
		'total_pago': total_pago,
		'total_recebido': total_recebido,
		'total_valor_imposto': total_valor_imposto,
		'total_valor_outros_custos': total_valor_outros_custos,
		'mostrar_pago': mostrar_pago,
		'mostrar_recebido': mostrar_recebido,
		'empresa_nome': (current_user.empresa.nome if current_user.empresa else '-'),
		'empresa_cnpj': (current_user.empresa.cnpj if current_user.empresa else '-'),
		'data_lanc_de': data_lanc_de,
		'data_lanc_ate': data_lanc_ate,
		'data_venc_de': data_venc_de,
		'data_venc_ate': data_venc_ate,
		'tipo_movimento': tipo_movimento,
		'entidade_tipo': entidade_tipo,
		'entidade_id': entidade_id,
		'conta_fluxo_id': conta_fluxo_id,
		'status': status,
		'gerado_em': datetime.now().strftime('%d/%m/%Y %H:%M')
	}

# --- LISTAGEM DE LANÇAMENTOS ---
@relatorios_bp.route('/lancamentos', methods=['GET'])
@login_required
def listagem_lancamentos():
	context = _build_listagem_lancamentos_context(request.args)
	return render_template('relatorios/listagem_lancamentos.html', **context)


@relatorios_bp.route('/lancamentos/export', methods=['GET'])
@login_required
def export_listagem_lancamentos():
	formato = request.args.get('formato', 'xlsx').lower()
	context = _build_listagem_lancamentos_context(request.args)
	lancamentos = context['lancamentos']

	if formato == 'xlsx':
		if Workbook is None:
			flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
			return redirect(url_for('relatorios.listagem_lancamentos', **request.args))

		wb = Workbook()
		ws = wb.active
		ws.title = 'Listagem Lancamentos'
		ws.append(['Empresa', context['empresa_nome']])
		ws.append(['CNPJ', context['empresa_cnpj']])
		ws.append(['Gerado em', context['gerado_em']])
		ws.append([])
		ws.append([
			'ID', 'Empresa', 'Processo', 'Status', 'Data Lancamento', 'Data Vencimento',
			'Data Pagamento', 'Entidade', 'Tipo Entidade', 'Conta Fluxo', 'Conta Banco',
			'Documento', 'Valor Real', 'Valor Pago/Recebido', 'Impostos', 'Outros Custos', 'Observacoes'
		])

		for l in lancamentos:
			processo = 'Pagamento (Saida)' if l.fluxo_conta and l.fluxo_conta.tipo == 'P' else 'Recebimento (Entrada)'
			tipo_entidade = l.entidade.get_tipo_descricao() if l.entidade else '-'
			conta_fluxo = f"{l.fluxo_conta.codigo} - {l.fluxo_conta.descricao}" if l.fluxo_conta else '-'
			ws.append([
				l.id,
				l.empresa.nome if l.empresa else '-',
				processo,
				l.status,
				l.data_evento.strftime('%d/%m/%Y') if l.data_evento else '-',
				l.data_vencimento.strftime('%d/%m/%Y') if l.data_vencimento else '-',
				l.data_pagamento.strftime('%d/%m/%Y') if l.data_pagamento else '-',
				l.entidade.nome if l.entidade else '-',
				tipo_entidade,
				conta_fluxo,
				l.conta_banco.nome if l.conta_banco else '-',
				l.numero_documento or '-',
				'',  # valor real removido
				float(l.valor_pago or 0) if l.fluxo_conta and l.fluxo_conta.tipo == 'P' else '',
				float(l.valor_pago or 0) if l.fluxo_conta and l.fluxo_conta.tipo == 'R' else '',
				float(l.valor_imposto or 0),
				float(l.valor_outros_custos or 0),
				l.observacoes or '-'
			])

		ws.append([])
		ws.append(['TOTAL', '', '', '', '', '', '', '', '', '', '', '', '', float(context['total_pago']), float(context['total_recebido']), float(context['total_valor_imposto']), float(context['total_valor_outros_custos']), ''])

		for col in ['N', 'O', 'P', 'Q']:
			for cell in ws[col]:
				cell.number_format = '#,##0.00'

		output = io.BytesIO()
		wb.save(output)
		output.seek(0)
		return send_file(
			output,
			as_attachment=True,
			download_name='listagem_lancamentos.xlsx',
			mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)

	if formato == 'pdf':
		try:
			from fpdf import FPDF
		except ImportError:
			flash('Exportação para PDF indisponível: biblioteca "fpdf" não está instalada no ambiente.', 'warning')
			return redirect(url_for('relatorios.listagem_lancamentos', **request.args))

		pdf = FPDF(orientation='L', unit='mm', format='A4')
		pdf.add_page()
		pdf.set_font('Arial', 'B', 11)
		pdf.cell(0, 8, 'Listagem de Lancamentos', ln=1)
		pdf.set_font('Arial', '', 9)
		pdf.cell(0, 6, f"Empresa: {context['empresa_nome']} | CNPJ: {context['empresa_cnpj']}", ln=1)
		pdf.cell(0, 6, f"Gerado em: {context['gerado_em']}", ln=1)
		pdf.ln(2)

		headers = [
			('ID', 10), ('Proc.', 20), ('Status', 18), ('Dt Lanc.', 22), ('Dt Venc.', 22),
			('Dt Pag.', 22), ('Entidade', 45), ('Conta Fluxo', 45), ('Documento', 24),
			('Vlr Real', 22), ('Vlr Pago', 22), ('Impostos', 22), ('Outros', 22)
		]

		pdf.set_font('Arial', 'B', 8)
		for titulo, largura in headers:
			pdf.cell(largura, 7, titulo, 1)
		pdf.ln()

		pdf.set_font('Arial', '', 7)
		for l in lancamentos:
			processo = 'Saida' if l.fluxo_conta and l.fluxo_conta.tipo == 'P' else 'Entrada'
			conta_fluxo = f"{l.fluxo_conta.codigo} - {l.fluxo_conta.descricao}" if l.fluxo_conta else '-'
			pdf.cell(10, 6, str(l.id), 1)
			pdf.cell(20, 6, processo, 1)
			pdf.cell(18, 6, str(l.status or '-'), 1)
			pdf.cell(22, 6, l.data_evento.strftime('%d/%m/%Y') if l.data_evento else '-', 1)
			pdf.cell(22, 6, l.data_vencimento.strftime('%d/%m/%Y') if l.data_vencimento else '-', 1)
			pdf.cell(22, 6, l.data_pagamento.strftime('%d/%m/%Y') if l.data_pagamento else '-', 1)
			pdf.cell(45, 6, (l.entidade.nome if l.entidade else '-')[:26], 1)
			pdf.cell(45, 6, conta_fluxo[:28], 1)
			pdf.cell(24, 6, (l.numero_documento or '-')[:14], 1)
			pdf.cell(22, 6, f"{float(l.valor_real or 0):,.2f}", 1, align='R')
			pdf.cell(22, 6, f"{float(l.valor_pago or 0):,.2f}", 1, align='R')
			pdf.cell(22, 6, f"{float(l.valor_imposto or 0):,.2f}", 1, align='R')
			pdf.cell(22, 6, f"{float(l.valor_outros_custos or 0):,.2f}", 1, align='R')
			pdf.ln()

		pdf.set_font('Arial', 'B', 8)
		pdf.cell(228, 7, 'TOTAIS', 1)
		pdf.cell(22, 7, f"{float(context['total_valor_real']):,.2f}", 1, align='R')
		pdf.cell(22, 7, f"{float(context['total_valor_pago']):,.2f}", 1, align='R')
		pdf.cell(22, 7, f"{float(context['total_valor_imposto']):,.2f}", 1, align='R')
		pdf.cell(22, 7, f"{float(context['total_valor_outros_custos']):,.2f}", 1, align='R')

		output = io.BytesIO()
		output.write(pdf.output(dest='S').encode('latin1'))
		output.seek(0)
		return send_file(output, as_attachment=True, download_name='listagem_lancamentos.pdf', mimetype='application/pdf')

	flash('Formato de exportação inválido.', 'warning')
	return redirect(url_for('relatorios.listagem_lancamentos', **request.args))
# Relatório de Fluxo de Caixa CSV
@relatorios_bp.route('/fluxo-caixa-csv')
@login_required
def fluxo_caixa_csv():
	data_inicio = request.args.get('data_inicio', '')
	data_fim = request.args.get('data_fim', '')
	# Filtros básicos
	query = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
	if data_inicio:
		data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
		query = query.filter(or_(Lancamento.data_pagamento >= data_inicio_dt, Lancamento.data_vencimento >= data_inicio_dt))
	if data_fim:
		data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
		query = query.filter(or_(Lancamento.data_pagamento <= data_fim_dt, Lancamento.data_vencimento <= data_fim_dt))
	# Apenas lançamentos da empresa do usuário
	if hasattr(current_user, 'empresa_id'):
		query = query.filter(Lancamento.empresa_id == current_user.empresa_id)
	query = query.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
	lancamentos = query.order_by(
		FluxoContaModel.descricao.asc(),
		func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc(),
		Lancamento.id.asc()
	).all()
	dados_csv = []
	for l in lancamentos:
		desc = getattr(l, 'descricao', None)
		if not desc:
			desc = l.observacoes or l.numero_documento or '-'
		conta_fluxo = '-'
		if l.fluxo_conta:
			codigo = getattr(l.fluxo_conta, 'codigo', None)
			descricao = getattr(l.fluxo_conta, 'descricao', None)
			if codigo and descricao:
				conta_fluxo = f"{codigo} - {descricao}"
			elif descricao:
				conta_fluxo = descricao
			elif codigo:
				conta_fluxo = codigo
		# Data preferencial: data_pagamento se existir, senão data_vencimento
		data_display = l.data_pagamento or l.data_vencimento
		dados_csv.append({
			'data': data_display.strftime('%d/%m/%Y') if data_display else '-',
			'descricao': desc,
			'conta_fluxo': conta_fluxo,
			'tipo': 'Receita' if l.fluxo_conta and l.fluxo_conta.tipo == 'R' else 'Despesa',
			'valor_real': l.valor_real or 0,
			'valor_pago': l.valor_pago or 0
		})
	return render_template('relatorios/fluxo_caixa_csv.html', dados_csv=dados_csv, data_inicio=data_inicio, data_fim=data_fim)

# Exportação para Excel do Fluxo de Caixa CSV
@relatorios_bp.route('/fluxo-caixa-csv/export')
@login_required
def export_fluxo_caixa_csv():
	data_inicio = request.args.get('data_inicio', '')
	data_fim = request.args.get('data_fim', '')
	# Se openpyxl não estiver disponível, retornar mensagem amigável
	if Workbook is None:
		flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
		return redirect(url_for('relatorios.fluxo_caixa'))
	query = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
	if data_inicio:
		data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
		query = query.filter(or_(Lancamento.data_pagamento >= data_inicio_dt, Lancamento.data_vencimento >= data_inicio_dt))
	if data_fim:
		data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
		query = query.filter(or_(Lancamento.data_pagamento <= data_fim_dt, Lancamento.data_vencimento <= data_fim_dt))
	if hasattr(current_user, 'empresa_id'):
		query = query.filter(Lancamento.empresa_id == current_user.empresa_id)
	query = query.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
	lancamentos = query.order_by(
		FluxoContaModel.descricao.asc(),
		func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc(),
		Lancamento.id.asc()
	).all()
	wb = Workbook()
	ws = wb.active
	ws.title = 'Fluxo de Caixa'
	ws.append(['Data', 'Conta Fluxo', 'Descrição', 'Tipo', 'Valor Real (R$)', 'Valor Pago/Recebido (R$)'])
	for l in lancamentos:
		valor_real_brl = f'R$ {(l.valor_real or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
		valor_pago_brl = f'R$ {(l.valor_pago or 0):,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
		conta_fluxo = '-'
		if l.fluxo_conta:
			if l.fluxo_conta.codigo and l.fluxo_conta.descricao:
				conta_fluxo = f"{l.fluxo_conta.codigo} - {l.fluxo_conta.descricao}"
			else:
				conta_fluxo = l.fluxo_conta.descricao or l.fluxo_conta.codigo or '-'
		ws.append([
			l.data_pagamento.strftime('%d/%m/%Y') if l.data_pagamento else '-',
			conta_fluxo,
			(getattr(l, 'descricao', None) or l.observacoes or l.numero_documento or '-'),
			'Receita' if l.fluxo_conta and l.fluxo_conta.tipo == 'R' else 'Despesa',
			valor_real_brl,
			valor_pago_brl
		])
	output = io.BytesIO()
	wb.save(output)
	output.seek(0)
	return send_file(output, as_attachment=True, download_name='fluxo_caixa.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@relatorios_bp.route('/fluxo-caixa-previsto')
@login_required
def fluxo_caixa_previsto():
	return render_template('relatorios/fluxo_caixa_previsto.html')

@relatorios_bp.route('/fluxo-caixa-realizado')
@login_required
def fluxo_caixa_realizado():
	return render_template('relatorios/fluxo_caixa_realizado.html')

@relatorios_bp.route('/fluxo-caixa')
@login_required
def fluxo_caixa():
	# Get filters
	data_inicio = request.args.get('data_inicio', '')
	data_fim = request.args.get('data_fim', '')
	conta_banco_id = request.args.get('conta_banco_id', '', type=int)
	conta_fluxo_id = request.args.get('conta_fluxo_id', '', type=int)

	def get_saldo_inicial_por_conta():
		if conta_banco_id:
			contas = ContaBanco.query.filter(
				ContaBanco.empresa_id == current_user.empresa_id,
				ContaBanco.id == conta_banco_id
			).all()
		else:
			contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
		return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

	def get_saldo_inicial_total():
		return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

	def build_fluxo_rows(lancamentos, saldo_inicial_por_conta, use_valor_real):
		saldo_atual_por_conta = saldo_inicial_por_conta.copy()
		rows = []
		for lancamento in lancamentos:
			conta_id = lancamento.conta_banco_id
			saldo_anterior = saldo_atual_por_conta.get(conta_id, Decimal('0.00'))
			valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
			valor = Decimal(str(valor_base or 0))
			if lancamento.fluxo_conta and lancamento.fluxo_conta.tipo == 'P':
				saldo_atual = saldo_anterior - valor
			else:
				saldo_atual = saldo_anterior + valor
			saldo_atual_por_conta[conta_id] = saldo_atual
			rows.append(SimpleNamespace(
				lancamento=lancamento,
				saldo_anterior=saldo_anterior,
				saldo_atual=saldo_atual
			))
		return rows

	def build_daily_rows(lancamentos, saldo_inicial, use_valor_real, date_attr):
		totals = defaultdict(lambda: {'pagar': Decimal('0.00'), 'receber': Decimal('0.00')})
		for lancamento in lancamentos:
			data_ref = getattr(lancamento, date_attr)
			if not data_ref:
				continue
			valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
			valor = Decimal(str(valor_base or 0))
			if lancamento.fluxo_conta and lancamento.fluxo_conta.tipo == 'P':
				totals[data_ref]['pagar'] += valor
			else:
				totals[data_ref]['receber'] += valor

		# Aqui você pode montar a lista de linhas diárias, exemplo:
		rows = []
		saldo_anterior = saldo_inicial
		for data in sorted(totals.keys()):
			pagamentos = totals[data]['pagar']
			recebimentos = totals[data]['receber']
			saldo_atual = saldo_anterior - pagamentos + recebimentos
			rows.append(SimpleNamespace(
				data=data,
				saldo_anterior=saldo_anterior,
				pagamentos=pagamentos,
				recebimentos=recebimentos,
				saldo_atual=saldo_atual
			))
			saldo_anterior = saldo_atual
		return rows

	saldo_inicial_por_conta = get_saldo_inicial_por_conta()
	saldo_inicial_total = get_saldo_inicial_total()

	# Build base queries
	query_realizado = Lancamento.query.filter(
		Lancamento.empresa_id == current_user.empresa_id,
		Lancamento.status == 'pago'
	)
	if data_inicio:
		data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
		query_realizado = query_realizado.filter(Lancamento.data_pagamento >= data_inicio)
	if data_fim:
		data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
		query_realizado = query_realizado.filter(Lancamento.data_pagamento <= data_fim)
	if conta_banco_id:
		query_realizado = query_realizado.filter(Lancamento.conta_banco_id == conta_banco_id)
	if conta_fluxo_id:
		query_realizado = query_realizado.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
	query_realizado = query_realizado.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
	lancamentos_realizado = query_realizado.order_by(
		FluxoContaModel.descricao.asc(),
		Lancamento.data_pagamento.asc(),
		Lancamento.id.asc()
	).all()

	# Build previsto query
	query_previsto = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)
	if data_inicio:
		query_previsto = query_previsto.filter(Lancamento.data_vencimento >= data_inicio)
	if data_fim:
		query_previsto = query_previsto.filter(Lancamento.data_vencimento <= data_fim)
	if conta_banco_id:
		query_previsto = query_previsto.filter(Lancamento.conta_banco_id == conta_banco_id)
	if conta_fluxo_id:
		query_previsto = query_previsto.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
	query_previsto = query_previsto.outerjoin(FluxoContaModel, Lancamento.fluxo_conta_id == FluxoContaModel.id)
	lancamentos_previsto = query_previsto.order_by(
		FluxoContaModel.descricao.asc(),
		Lancamento.data_vencimento.asc(),
		Lancamento.id.asc()
	).all()

	resumo_diario_realizado = build_daily_rows(
		lancamentos_realizado,
		saldo_inicial_total,
		use_valor_real=False,
		date_attr='data_pagamento'
	)

	resumo_diario_previsto = build_daily_rows(
		lancamentos_previsto,
		saldo_inicial_total,
		use_valor_real=True,
		date_attr='data_vencimento'
	)

	# Get filter options
	contas_banco = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
	contas_fluxo = FluxoContaModel.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()

	return render_template(
		'relatorios/fluxo_caixa.html',
		lancamentos_realizado=build_fluxo_rows(lancamentos_realizado, saldo_inicial_por_conta, use_valor_real=False),
		lancamentos_previsto=build_fluxo_rows(lancamentos_previsto, saldo_inicial_por_conta, use_valor_real=True),
		resumo_diario_realizado=resumo_diario_realizado,
		resumo_diario_previsto=resumo_diario_previsto,
		contas_banco=contas_banco,
		contas_fluxo=contas_fluxo,
		data_inicio=data_inicio,
		data_fim=data_fim,
		conta_banco_id=conta_banco_id,
		conta_fluxo_id=conta_fluxo_id
	)

@relatorios_bp.route('/fluxo-caixa/export')
@login_required
def export_fluxo_caixa():
	data_inicio = request.args.get('data_inicio', '')
	data_fim = request.args.get('data_fim', '')
	conta_banco_id = request.args.get('conta_banco_id', '', type=int)
	conta_fluxo_id = request.args.get('conta_fluxo_id', '', type=int)
	# Se openpyxl não estiver disponível, retornar mensagem amigável
	if Workbook is None:
		flash('Exportação para Excel indisponível: biblioteca "openpyxl" não está instalada no ambiente.', 'warning')
		return redirect(url_for('relatorios.fluxo_caixa'))

	def get_saldo_inicial_por_conta():
		if conta_banco_id:
			contas = ContaBanco.query.filter(
				ContaBanco.empresa_id == current_user.empresa_id,
				ContaBanco.id == conta_banco_id
			).all()
		else:
			contas = ContaBanco.query.filter_by(empresa_id=current_user.empresa_id, ativo=True).all()
		return {c.id: Decimal(str(c.saldo_inicial or 0)) for c in contas}

	def get_saldo_inicial_total():
		return sum(get_saldo_inicial_por_conta().values(), Decimal('0.00'))

	def build_daily_rows(lancamentos, saldo_inicial, use_valor_real, date_attr):
		totals = defaultdict(lambda: {'pagar': Decimal('0.00'), 'receber': Decimal('0.00')})
		for lancamento in lancamentos:
			data_ref = getattr(lancamento, date_attr)
			if not data_ref:
				continue
			valor_base = lancamento.valor_real if use_valor_real else lancamento.valor_pago
			valor = Decimal(str(valor_base or 0))
			if lancamento.fluxo_conta and lancamento.fluxo_conta.tipo == 'P':
				totals[data_ref]['pagar'] += valor
			else:
				totals[data_ref]['receber'] += valor

		saldo_atual = saldo_inicial
		rows = []
		for data_ref in sorted(totals.keys()):
			pagar = totals[data_ref]['pagar']
			receber = totals[data_ref]['receber']
			saldo_anterior = saldo_atual
			saldo_atual = saldo_anterior + receber - pagar
			rows.append((data_ref, saldo_anterior, pagar, receber, saldo_atual))
		return rows

	query_realizado = Lancamento.query.filter(
		Lancamento.empresa_id == current_user.empresa_id,
		Lancamento.status == 'pago'
	)
	query_previsto = Lancamento.query.filter_by(empresa_id=current_user.empresa_id)

	if data_inicio:
		data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
		query_realizado = query_realizado.filter(Lancamento.data_pagamento >= data_inicio)
		query_previsto = query_previsto.filter(Lancamento.data_vencimento >= data_inicio)

	if data_fim:
		data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
		query_realizado = query_realizado.filter(Lancamento.data_pagamento <= data_fim)
		query_previsto = query_previsto.filter(Lancamento.data_vencimento <= data_fim)

	if conta_banco_id:
		query_realizado = query_realizado.filter(Lancamento.conta_banco_id == conta_banco_id)
		query_previsto = query_previsto.filter(Lancamento.conta_banco_id == conta_banco_id)

	if conta_fluxo_id:
		query_realizado = query_realizado.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)
		query_previsto = query_previsto.filter(Lancamento.fluxo_conta_id == conta_fluxo_id)

	lancamentos_realizado = query_realizado.order_by(func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc()).all()
	lancamentos_previsto = query_previsto.order_by(func.coalesce(Lancamento.data_pagamento, Lancamento.data_vencimento).asc()).all()

	saldo_inicial_total = get_saldo_inicial_total()
	resumo_realizado = build_daily_rows(lancamentos_realizado, saldo_inicial_total, False, 'data_pagamento')
	resumo_previsto = build_daily_rows(lancamentos_previsto, saldo_inicial_total, True, 'data_vencimento')

	wb = Workbook()
	ws_previsto = wb.active
	ws_previsto.title = 'Previsto'
	ws_realizado = wb.create_sheet('Realizado')

	headers = ['Data', 'Saldo Anterior', 'Pagamentos', 'Recebimentos', 'Saldo do Dia']
	ws_previsto.append(headers)
	ws_realizado.append(headers)

	for data_ref, saldo_anterior, pagar, receber, saldo_atual in resumo_previsto:
		ws_previsto.append([data_ref.strftime('%d/%m/%Y'), float(saldo_anterior), float(pagar), float(receber), float(saldo_atual)])

	for data_ref, saldo_anterior, pagar, receber, saldo_atual in resumo_realizado:
		ws_realizado.append([data_ref.strftime('%d/%m/%Y'), float(saldo_anterior), float(pagar), float(receber), float(saldo_atual)])

	for sheet in (ws_previsto, ws_realizado):
		for col in ['B', 'C', 'D', 'E']:
			for cell in sheet[col]:
				cell.number_format = '#,##0.00'
		sheet.column_dimensions['A'].width = 14
		sheet.column_dimensions['B'].width = 18
		sheet.column_dimensions['C'].width = 16
		sheet.column_dimensions['D'].width = 16
		sheet.column_dimensions['E'].width = 18

	output = io.BytesIO()
	wb.save(output)
	output.seek(0)

	return send_file(
		output,
		as_attachment=True,
		download_name='fluxo_caixa_diario.xlsx',
		mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	)

